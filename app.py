 from flask import Flask, request, jsonify, render_template_string
from flask_cors import CORS
import openpyxl
import json
import os
from datetime import datetime
import io
import gzip
import base64

app = Flask(__name__)
CORS(app)

# Archivo donde se guardan los datos
DATOS_FILE = 'datos.json'

def leer_excel_y_convertir(archivo_excel):
    """Convierte el Excel a formato JSON usando openpyxl - VERSIÓN OPTIMIZADA"""
    try:
        # Leer directamente desde memoria sin guardar archivo temporal
        archivo_excel.seek(0)  # Asegurar que estamos al inicio del archivo
        file_content = archivo_excel.read()
        
        # Crear objeto BytesIO para openpyxl
        excel_file = io.BytesIO(file_content)
        
        # Cargar workbook directamente desde memoria
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        datos = {
            'fecha_actualizacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'tfn': [],
            'tfn_cncaf': [],
            'tfn_cncaf_csjn': []
        }
        
        # Procesar cada hoja
        sheet_mapping = {
            'TFN': 'tfn',
            'TFN_CNCAF': 'tfn_cncaf', 
            'TFN_CNCAF_CSJN': 'tfn_cncaf_csjn'
        }
        
        for sheet_name, data_key in sheet_mapping.items():
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Verificar que la hoja no esté vacía
                if sheet.max_row < 2:
                    print(f"Advertencia: La hoja {sheet_name} está vacía o solo tiene headers")
                    continue
                
                # Leer headers (primera fila) - mejorado
                headers = []
                header_row = sheet[1]
                for i, cell in enumerate(header_row):
                    if cell.value is not None:
                        header_value = str(cell.value).strip()
                        if header_value:
                            headers.append(header_value)
                        else:
                            headers.append(f'columna_{i + 1}')
                    else:
                        headers.append(f'columna_{i + 1}')
                
                print(f"Headers encontrados en {sheet_name}: {headers}")
                
                # Leer datos (desde fila 2 en adelante)
                sheet_data = []
                processed_rows = 0
                
                for row_num in range(2, sheet.max_row + 1):
                    row = sheet[row_num]
                    
                    # Verificar que la fila no esté completamente vacía
                    row_values = [cell.value for cell in row]
                    if not any(val is not None and str(val).strip() != '' for val in row_values):
                        continue
                    
                    row_dict = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            value = cell.value
                            if value is None:
                                cleaned_value = ''
                            else:
                                cleaned_value = str(value).strip()
                                # Limpiar y truncar texto muy largo para evitar problemas
                                cleaned_value = cleaned_value.replace('\r\n', ' ').replace('\n', ' ')
                                # Truncar si es muy largo (más de 10000 caracteres)
                                if len(cleaned_value) > 10000:
                                    cleaned_value = cleaned_value[:10000] + "... [TRUNCADO]"
                            
                            row_dict[headers[i]] = cleaned_value
                    
                    # Solo agregar si hay al menos un valor no vacío
                    if any(val.strip() for val in row_dict.values() if val):
                        sheet_data.append(row_dict)
                        processed_rows += 1
                
                datos[data_key] = sheet_data
                print(f"Procesados {processed_rows} registros de {sheet_name}")
        
        # Cerrar el workbook
        workbook.close()
        
        return datos
        
    except Exception as e:
        print(f"Error detallado procesando Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Error procesando Excel: {str(e)}")

def comprimir_datos(datos):
    """Comprime los datos usando gzip"""
    json_str = json.dumps(datos, ensure_ascii=False, separators=(',', ':'))
    compressed = gzip.compress(json_str.encode('utf-8'))
    return base64.b64encode(compressed).decode('ascii')

def descomprimir_datos(compressed_data):
    """Descomprime los datos"""
    compressed_bytes = base64.b64decode(compressed_data.encode('ascii'))
    json_str = gzip.decompress(compressed_bytes).decode('utf-8')
    return json.loads(json_str)

@app.route('/')
def home():
    return "Backend del Boletín de Trazabilidad funcionando correctamente"

@app.route('/admin')
def admin():
    """Página simple para subir archivos"""
    return render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Admin - Boletín de Trazabilidad</title>
        <meta charset="UTF-8">
        <style>
            body { font-family: Arial, sans-serif; max-width: 800px; margin: 50px auto; padding: 20px; }
            .upload-area { border: 2px dashed #ccc; padding: 40px; text-align: center; margin: 20px 0; border-radius: 10px; }
            button { background: #007cba; color: white; padding: 12px 24px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px; }
            button:hover { background: #005a8b; }
            .status { margin: 20px 0; padding: 15px; border-radius: 5px; }
            .success { background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }
            .error { background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }
            .info { background: #e2e3e5; color: #383d41; border: 1px solid #d6d8db; }
            .debug { background: #fff3cd; color: #856404; border: 1px solid #ffeaa7; font-family: monospace; font-size: 12px; white-space: pre-wrap; }
            .json-preview { background: #f8f9fa; border: 1px solid #dee2e6; padding: 15px; border-radius: 5px; font-family: monospace; font-size: 12px; max-height: 300px; overflow-y: auto; }
        </style>
    </head>
    <body>
        <h1>Panel de Administración</h1>
        <h2>Boletín de Trazabilidad de Sentencias</h2>
        
        <div class="upload-area">
            <h3>Subir archivo Excel</h3>
            <p>Selecciona el archivo Excel con las 3 hojas: TFN, TFN_CNCAF, TFN_CNCAF_CSJN</p>
            <input type="file" id="fileInput" accept=".xlsx,.xls" style="margin: 10px;">
            <br><br>
            <button onclick="subirArchivo()">Actualizar Boletín</button>
        </div>
        
        <div id="status"></div>
        
        <div style="margin-top: 30px;">
            <h3>Enlaces útiles:</h3>
            <p><a href="/api/datos" target="_blank">Ver datos JSON</a></p>
            <p><a href="/api/datos/compressed" target="_blank">Ver datos comprimidos</a></p>
            <p><a href="/api/debug" target="_blank">Debug info</a></p>
            <p><a href="/api/stats" target="_blank">Estadísticas</a></p>
        </div>
        
        <script>
            function subirArchivo() {
                const fileInput = document.getElementById('fileInput');
                const statusDiv = document.getElementById('status');
                
                if (!fileInput.files[0]) {
                    statusDiv.innerHTML = '<div class="error">Por favor selecciona un archivo</div>';
                    return;
                }
                
                const formData = new FormData();
                formData.append('archivo', fileInput.files[0]);
                
                statusDiv.innerHTML = '<div class="info">Procesando archivo...</div>';
                
                fetch('/api/subir', {
                    method: 'POST',
                    body: formData
                })
                .then(response => response.json())
                .then(data => {
                    if (data.error) {
                        statusDiv.innerHTML = `<div class="error">Error: ${data.error}</div>`;
                    } else {
                        let html = `<div class="success">
                            ✓ Archivo procesado exitosamente<br>
                            Fecha: ${data.fecha_actualizacion}<br>
                            TFN: ${data.total_tfn} registros<br>
                            TFN-CNCAF: ${data.total_tfn_cncaf} registros<br>
                            TFN-CNCAF-CSJN: ${data.total_tfn_cncaf_csjn} registros<br>
                            Tamaño del archivo: ${data.file_size_mb} MB
                        </div>`;
                        
                        if (data.sample_data) {
                            html += `<div class="debug">Muestra de datos procesados:<br>${JSON.stringify(data.sample_data, null, 2)}</div>`;
                        }
                        
                        statusDiv.innerHTML = html;
                        fileInput.value = '';
                    }
                })
                .catch(error => {
                    statusDiv.innerHTML = `<div class="error">Error de conexión: ${error}</div>`;
                });
            }
        </script>
    </body>
    </html>
    ''')

@app.route('/api/subir', methods=['POST'])
def subir_archivo():
    """Endpoint para subir y procesar el Excel"""
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Solo se permiten archivos Excel (.xlsx, .xls)'}), 400
        
        print(f"Procesando archivo: {archivo.filename}")
        
        # Procesar Excel
        datos = leer_excel_y_convertir(archivo)
        
        # Calcular tamaño del JSON
        json_str = json.dumps(datos, ensure_ascii=False, indent=2)
        json_size_mb = len(json_str.encode('utf-8')) / (1024 * 1024)
        
        print(f"Tamaño del JSON: {json_size_mb:.2f} MB")
        
        # Guardar datos normales
        with open(DATOS_FILE, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        
        # Guardar versión comprimida también
        compressed_data = comprimir_datos(datos)
        with open(DATOS_FILE + '.compressed', 'w', encoding='utf-8') as f:
            f.write(compressed_data)
        
        print("Archivos guardados exitosamente")
        
        # Crear muestra de datos para debug
        sample_data = {}
        for key, value in datos.items():
            if key != 'fecha_actualizacion' and isinstance(value, list) and len(value) > 0:
                sample_data[key] = {
                    'primer_registro': value[0] if value else None,
                    'total_registros': len(value)
                }
        
        # Respuesta con estadísticas
        response_data = {
            'mensaje': 'Archivo procesado exitosamente',
            'fecha_actualizacion': datos['fecha_actualizacion'],
            'total_tfn': len(datos['tfn']),
            'total_tfn_cncaf': len(datos['tfn_cncaf']),
            'total_tfn_cncaf_csjn': len(datos['tfn_cncaf_csjn']),
            'file_size_mb': round(json_size_mb, 2),
            'sample_data': sample_data
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error en subir_archivo: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/datos')
def obtener_datos():
    """Endpoint que devuelve los datos para el frontend"""
    try:
        if not os.path.exists(DATOS_FILE):
            return jsonify({'error': 'No hay datos disponibles. Sube un archivo Excel primero.'}), 404
        
        with open(DATOS_FILE, 'r', encoding='utf-8') as f:
            datos = json.load(f)
        
        return jsonify(datos)
        
    except Exception as e:
        print(f"Error en obtener_datos: {str(e)}")
        return jsonify({'error': f'Error cargando datos: {str(e)}'}), 500

@app.route('/api/datos/compressed')
def obtener_datos_comprimidos():
    """Endpoint que devuelve los datos comprimidos"""
    try:
        compressed_file = DATOS_FILE + '.compressed'
        if not os.path.exists(compressed_file):
            return jsonify({'error': 'No hay datos comprimidos disponibles.'}), 404
        
        with open(compressed_file, 'r', encoding='utf-8') as f:
            compressed_data = f.read()
        
        datos = descomprimir_datos(compressed_data)
        return jsonify(datos)
        
    except Exception as e:
        print(f"Error en obtener_datos_comprimidos: {str(e)}")
        return jsonify({'error': f'Error cargando datos comprimidos: {str(e)}'}), 500

@app.route('/api/stats')
def obtener_estadisticas():
    """Endpoint que devuelve solo estadísticas sin los datos completos"""
    try:
        if not os.path.exists(DATOS_FILE):
            return jsonify({'error': 'No hay datos disponibles.'}), 404
        
        with open(DATOS_FILE, 'r', encoding='utf-8') as f:
            datos = json.load(f)
        
        # Solo devolver estadísticas y primeros registros
        stats = {
            'fecha_actualizacion': datos.get('fecha_actualizacion'),
            'total_tfn': len(datos.get('tfn', [])),
            'total_tfn_cncaf': len(datos.get('tfn_cncaf', [])),
            'total_tfn_cncaf_csjn': len(datos.get('tfn_cncaf_csjn', [])),
            'muestra_tfn': datos.get('tfn', [])[:2],  # Solo primeros 2 registros
            'muestra_tfn_cncaf': datos.get('tfn_cncaf', [])[:2],
            'muestra_tfn_cncaf_csjn': datos.get('tfn_cncaf_csjn', [])[:2]
        }
        
        return jsonify(stats)
        
    except Exception as e:
        return jsonify({'error': f'Error cargando estadísticas: {str(e)}'}), 500

@app.route('/api/debug')
def debug_info():
    """Endpoint para debugging"""
    info = {
        'archivo_existe': os.path.exists(DATOS_FILE),
        'archivo_comprimido_existe': os.path.exists(DATOS_FILE + '.compressed'),
        'directorio_actual': os.getcwd(),
        'archivos_en_directorio': os.listdir('.'),
        'python_version': os.sys.version,
        'render_environment': 'RENDER' in os.environ
    }
    
    if os.path.exists(DATOS_FILE):
        try:
            file_size = os.path.getsize(DATOS_FILE)
            info['archivo_size_mb'] = round(file_size / (1024 * 1024), 2)
            
            with open(DATOS_FILE, 'r', encoding='utf-8') as f:
                datos = json.load(f)
            info['datos_stats'] = {
                'fecha_actualizacion': datos.get('fecha_actualizacion', 'N/A'),
                'tfn_count': len(datos.get('tfn', [])),
                'tfn_cncaf_count': len(datos.get('tfn_cncaf', [])),
                'tfn_cncaf_csjn_count': len(datos.get('tfn_cncaf_csjn', []))
            }
        except Exception as e:
            info['error_leyendo_datos'] = str(e)
    
    return jsonify(info)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 10000))
    app.run(debug=False, host='0.0.0.0', port=port)
